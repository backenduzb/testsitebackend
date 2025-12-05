from tests.models import TestCase, Test, Subject
from openpyxl.styles import Alignment
from accounts.models import User
from scores.models import Score
from openpyxl import Workbook
import os

wb = Workbook()

def make_xlsx( subject_name: str ):
    
    scores = Score.objects.filter(test__subject__name=subject_name).all()

    # print(scores)
    print("Succes  _______________________________________________")
    ws = wb.active
    ws.merge_cells("A1:G1")
    
    ws["A1"] = f"{subject_name} fanidan natijalar"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.column_dimensions["A"].width = 5
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"] = "NO"

    ws.column_dimensions["B"].width = 70
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"] = "F.I.SH" 

    ws.column_dimensions["C"].width = 9
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C2"] = "SINFI" 

    ws.column_dimensions["D"].width = 10
    ws["D2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["D2"] = "BILISH" 
    
    ws.column_dimensions["E"].width = 15
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["E2"] = "QO'LLASH"

    ws.column_dimensions["F"].width = 15
    ws["F2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["F2"] = "MUHOKAMA" 

    ws.column_dimensions["G"].width = 10
    ws["G2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["G2"] = "JAMI" 

    cl = 2
    n = 0
    for score in scores:
        n += 1
        cl += 1

        ws[f"A{cl}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{cl}"] = n

        ws[f"B{cl}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{cl}"] = score.user.full_name 

        ws[f"C{cl}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{cl}"] = score.user.class_name.name 

        ws[f"D{cl}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{cl}"] = score.bilish 
        
        ws[f"E{cl}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"E{cl}"] = score.qollash

        ws[f"F{cl}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{cl}"] = score.muhokama

        ws[f"G{cl}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"G{cl}"] = score.total 

    wb.save(f"{subject_name}, Natijalar.xlsx")
